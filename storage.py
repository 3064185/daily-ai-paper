"""
Excel and SQLite persistence for daily reports.
"""
import logging
import sqlite3
from datetime import date
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from config import OUTPUT_DIR

logger = logging.getLogger("storage")

# ── Database paths ────────────────────────────────────────────────────
DB_DIR = Path(__file__).parent
AIHOT_DB = DB_DIR / "aihot_database.sqlite"
PAPERS_DB = DB_DIR / "papers_database.sqlite"


def _today() -> date:
    from config import RUN_DATE
    return RUN_DATE or date.today()


###############################################################################
#  SQLite
###############################################################################
def _init_db(db_path: Path, schema: str):
    db_path.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(str(db_path))
    conn.execute(schema)
    conn.commit()
    return conn


def save_news_to_sqlite(news_items: list[dict]):
    """Save AIHOT news to SQLite."""
    conn = _init_db(AIHOT_DB, """
        CREATE TABLE IF NOT EXISTS news (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            date TEXT, title TEXT, link TEXT, summary TEXT, detail TEXT, source TEXT
        )
    """)
    today = _today().isoformat()
    for item in news_items:
        conn.execute(
            "INSERT OR IGNORE INTO news (date, title, link, summary, detail, source) VALUES (?,?,?,?,?,?)",
            (today, item.get("title"), item.get("link"), item.get("summary"),
             item.get("detail", ""), item.get("source", "AIHOT"))
        )
    conn.commit()
    conn.close()
    logger.info("Saved %d news to SQLite", len(news_items))


def save_papers_to_sqlite(papers: list[dict]):
    """Save papers to SQLite."""
    conn = _init_db(PAPERS_DB, """
        CREATE TABLE IF NOT EXISTS papers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            date TEXT, title TEXT, abstract TEXT, authors TEXT,
            url TEXT, pdf_url TEXT, source TEXT, venue TEXT,
            relevance_score INTEGER, innovation_type TEXT,
            key_contribution TEXT, key_method TEXT,
            cited_by INTEGER, published TEXT
        )
    """)
    today = _today().isoformat()
    for p in papers:
        conn.execute(
            """INSERT OR IGNORE INTO papers
               (date, title, abstract, authors, url, pdf_url, source, venue,
                relevance_score, innovation_type, key_contribution, key_method,
                cited_by, published)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (today, p.get("title"), p.get("abstract"), ",".join(p.get("authors", [])),
             p.get("url"), p.get("pdf_url"), p.get("source"), p.get("venue", ""),
             p.get("relevance_score", 0), p.get("innovation_type", ""),
             p.get("key_contribution", ""), p.get("key_method", ""),
             p.get("cited_by", 0), p.get("published", ""))
        )
    conn.commit()
    conn.close()
    logger.info("Saved %d papers to SQLite", len(papers))


###############################################################################
#  Excel
###############################################################################
def _header_style():
    return {
        "font": Font(bold=True, color="FFFFFF", size=11),
        "fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
        "alignment": Alignment(horizontal="center", vertical="center"),
    }


def save_news_to_excel(news_items: list[dict]) -> Optional[Path]:
    """Save news to Excel."""
    if not news_items:
        return None
    today = _today()
    path = OUTPUT_DIR / f"daily_aihot_{today.strftime('%Y%m%d')}.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "AIHOT 新闻"
    headers = ["序号", "标题", "摘要", "日期", "链接"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = _header_style()["font"]
        cell.fill = _header_style()["fill"]
        cell.alignment = _header_style()["alignment"]
    for i, item in enumerate(news_items[:50], 1):
        ws.cell(row=i+1, column=1, value=i)
        ws.cell(row=i+1, column=2, value=item.get("title", ""))
        ws.cell(row=i+1, column=3, value=item.get("summary", "")[:200])
        ws.cell(row=i+1, column=4, value=item.get("date", ""))
        ws.cell(row=i+1, column=5, value=item.get("link", ""))
    # Column widths
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 60
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 50
    wb.save(str(path))
    logger.info("News Excel saved: %s", path.name)
    return path


def save_papers_to_excel(papers: list[dict]) -> Optional[Path]:
    """Save papers to Excel."""
    if not papers:
        return None
    today = _today()
    path = OUTPUT_DIR / f"daily_cs_papers_{today.strftime('%Y%m%d')}.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "CS 论文"
    headers = ["序号", "标题", "来源", "相关性", "创新类型", "核心贡献", "方法",
               "发表处", "被引", "作者", "链接"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = _header_style()["font"]
        cell.fill = _header_style()["fill"]
        cell.alignment = _header_style()["alignment"]
    for i, p in enumerate(papers[:100], 1):
        ws.cell(row=i+1, column=1, value=i)
        ws.cell(row=i+1, column=2, value=p.get("title", ""))
        ws.cell(row=i+1, column=3, value=p.get("source", ""))
        ws.cell(row=i+1, column=4, value=p.get("relevance_score", ""))
        ws.cell(row=i+1, column=5, value=p.get("innovation_type", ""))
        ws.cell(row=i+1, column=6, value=p.get("key_contribution", "")[:200])
        ws.cell(row=i+1, column=7, value=p.get("key_method", "")[:200])
        ws.cell(row=i+1, column=8, value=p.get("venue", ""))
        ws.cell(row=i+1, column=9, value=p.get("cited_by", ""))
        ws.cell(row=i+1, column=10, value=", ".join(p.get("authors", [])[:3]))
        ws.cell(row=i+1, column=11, value=p.get("url", ""))
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 40
    ws.column_dimensions["G"].width = 40
    ws.column_dimensions["H"].width = 20
    ws.column_dimensions["I"].width = 8
    ws.column_dimensions["J"].width = 30
    ws.column_dimensions["K"].width = 50
    wb.save(str(path))
    logger.info("Papers Excel saved: %s", path.name)
    return path
